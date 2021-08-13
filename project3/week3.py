from NaverNewsCrawler import NaverNewsCrawler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re
from openpyxl import Workbook, load_workbook
import json

####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요
crawler = NaverNewsCrawler(input("관련 기사를 수집할 키워드를 입력하세요. : "))

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요
filename = input("수집한 내용을 저장할 파일 이름을 입력하세요 : " ) + str ("xlsx")

# 확장자없이 입력할 수 있게, input으로 받은 파일 이름에 확장자를 붙여줌
crawler.get_news(filename)

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
# 보안상 문제를 보완하기 위해 다른 파일에서 import
with open('conf.json') as f:
    config = json.load(f)
SMTP_USER = config['id']
SMTP_PASSWORD = config['password']

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

## Openpyxl-Workbook을 통해 시트 생성
wb = Workbook()
ws = wb.create_sheet("info")
## 시트에 행별로 데이터 추가
ws.append(['번호', '이름', '이메일'])
ws.append(['1', '배프', 'beap.jg@gmail.com'])
ws.append(['2', '송중근', 'masterofflash@nate.com']) 
## 파일 저장
wb.save("email_list.xlsx")

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.
wb = load_workbook('email_list.xlsx')
data = wb['info']
contents = '''
크롤링 메일입니다.
'''

# 불러온 데이터를 반복문을 통해 함수의 parameter로 바로 삽입
# min_row를 2로 지정해서 번호, 이름, 이메일은 불러오지 않음
for row in data.iter_rows(min_row=2):
    # name = row[1].value
    # addr = row[2].value
    send_mail(row[1].value, row[2].value, '크롤링 결과입니다.' ,contents, attachment=filename+".xlsx")

        
    









